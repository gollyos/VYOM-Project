"""
VYOM Local Excel & Spreadsheet Generator
========================================
Creates, reads, and updates clean .xlsx and .csv spreadsheets on the local filesystem
based on exact user-defined columns, rows, and styles. Zero cloud dependency.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Sequence
from datetime import datetime


class LocalExcelService:
    """Zero-cloud local Excel/CSV generator and modifier for VYOM."""

    def __init__(self, default_output_dir: Path | None = None):
        self.output_dir = default_output_dir or Path("services/brain/data/artifacts")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def create_spreadsheet(
        self,
        filename: str,
        headers: Sequence[str],
        rows: Sequence[Sequence[Any]],
        *,
        sheet_name: str = "Sheet1",
        output_dir: Path | None = None,
        as_csv: bool = False,
    ) -> Path:
        """Create a new formatted Excel (.xlsx) or CSV file with exact headers and rows.

        Args:
            filename: Name of the file (e.g. 'leads_export.xlsx' or 'sales.csv')
            headers: List of column header names (e.g. ['Name', 'Company', 'Email'])
            rows: List of row tuples/lists
            sheet_name: Worksheet title
            output_dir: Destination folder (defaults to artifacts directory)
            as_csv: If True, writes standard CSV format.
        """
        dest_dir = output_dir or self.output_dir
        dest_dir.mkdir(parents=True, exist_ok=True)

        if not filename.endswith((".xlsx", ".csv")):
            filename = f"{filename}.csv" if as_csv else f"{filename}.xlsx"

        file_path = dest_dir / filename

        if filename.endswith(".csv") or as_csv:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for r in rows:
                    writer.writerow(r)
            return file_path

        # Write styled XLSX workbook
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header styling: Deep navy header with white bold text
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        border_side = Side(style="thin", color="D9D9D9")
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        ws.append(list(headers))
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = cell_border

        # Append data rows with clean formatting
        row_font = Font(name="Calibri", size=11)
        for r_idx, row_data in enumerate(rows, start=2):
            ws.append(list(row_data))
            # Zebra striping for readability
            bg_color = "F2F5F9" if r_idx % 2 == 0 else "FFFFFF"
            row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=r_idx, column=col_idx)
                c.font = row_font
                c.fill = row_fill
                c.border = cell_border
                # Right align numbers, left align text
                if isinstance(c.value, (int, float)):
                    c.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(str(file_path))
        return file_path

    def read_spreadsheet(self, file_path: Path | str) -> dict[str, Any]:
        """Read headers and rows from an existing .xlsx or .csv spreadsheet."""
        p = Path(file_path)
        if not p.exists():
            raise FileNotFoundError(f"Spreadsheet not found at: {p}")

        if p.suffix.lower() == ".csv":
            with open(p, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = list(reader)
                if not rows:
                    return {"headers": [], "rows": [], "row_count": 0}
                return {"headers": rows[0], "rows": rows[1:], "row_count": len(rows) - 1}

        import openpyxl
        wb = openpyxl.load_workbook(str(p), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return {"headers": [], "rows": [], "row_count": 0}
        headers = [str(c or "") for c in all_rows[0]]
        data_rows = [list(r) for r in all_rows[1:] if any(c is not None for c in r)]
        return {"headers": headers, "rows": data_rows, "row_count": len(data_rows)}


_default_service: LocalExcelService | None = None

def get_local_excel_service() -> LocalExcelService:
    global _default_service
    if _default_service is None:
        _default_service = LocalExcelService()
    return _default_service
